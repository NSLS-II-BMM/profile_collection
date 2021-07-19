

run_report(__file__, text='import the rest of the things')

run_report('\t'+'resting state')
from BMM.resting_state import resting_state, resting_state_plan, end_of_macro

run_report('\t'+'motor status reporting')
from BMM.motor_status import motor_metadata, motor_status, ms, motor_sidebar, xrd_motors, xrdm

run_report('\t'+'derived plot')
from BMM.derivedplot import close_all_plots, close_last_plot, interpret_click

run_report('\t'+'suspenders')
from BMM.suspenders import BMM_suspenders, BMM_clear_to_start, BMM_clear_suspenders

run_report('\t'+'linescan, rocking curve, slit_height, pluck')
from BMM.linescans import linescan, pluck, rocking_curve, slit_height, ls2dat

run_report('\t'+'positioning of instruments')
from BMM.positioning import find_slot

###########################################################################################
#  _____  _       ___   _   _ _____ _____ _   _ _____    ___   _   _ _____  _      _____  #
# |  __ \| |     / _ \ | \ | /  __ \_   _| \ | |  __ \  / _ \ | \ | |  __ \| |    |  ___| #
# | |  \/| |    / /_\ \|  \| | /  \/ | | |  \| | |  \/ / /_\ \|  \| | |  \/| |    | |__   #
# | | __ | |    |  _  || . ` | |     | | | . ` | | __  |  _  || . ` | | __ | |    |  __|  #
# | |_\ \| |____| | | || |\  | \__/\_| |_| |\  | |_\ \ | | | || |\  | |_\ \| |____| |___  #
#  \____/\_____/\_| |_/\_| \_/\____/\___/\_| \_/\____/ \_| |_/\_| \_/\____/\_____/\____/  #
###########################################################################################


run_report('\tglancing angle stage')
from BMM.glancing_angle import GlancingAngle, PinWheelMacroBuilder
ga = GlancingAngle('XF:06BMB-CT{DIODE-Local:1}', name='glancing angle stage')

pinwheel = PinWheelMacroBuilder()

#########################################################################################
# ___  ___  ___  _____ ______ _____  ______ _   _ _____ _    ______ ___________  _____  #
# |  \/  | / _ \/  __ \| ___ \  _  | | ___ \ | | |_   _| |   |  _  \  ___| ___ \/  ___| #
# | .  . |/ /_\ \ /  \/| |_/ / | | | | |_/ / | | | | | | |   | | | | |__ | |_/ /\ `--.  #
# | |\/| ||  _  | |    |    /| | | | | ___ \ | | | | | | |   | | | |  __||    /  `--. \ #
# | |  | || | | | \__/\| |\ \\ \_/ / | |_/ / |_| |_| |_| |___| |/ /| |___| |\ \ /\__/ / #
# \_|  |_/\_| |_/\____/\_| \_|\___/  \____/ \___/ \___/\_____/___/ \____/\_| \_|\____/  #
#########################################################################################

from BMM.functions import present_options, bold_msg
from openpyxl import load_workbook

def xlsx():
    '''Prompt for a macro building spreadsheet for any instrument. Use the
    content of cell B1 to direct this spreadsheet to the correct builder.

    if cell B1 is "Glancing angle" --> build a glancing angle macro

    if cell B1 is "Sample wheel" --> build a sample wheel macro

    if cell B1 is empty --> build a sample wheel macro

    Then prompt for the sheet, if there are more than 1 sheet in the
    spreadsheet file.
    
    '''
    spreadsheet = present_options('xlsx')
    if spreadsheet is None:
        print(error_msg('No spreadsheet specified!'))
        return None

    wb = load_workbook(os.path.join(BMMuser.folder, spreadsheet), read_only=True);
    #ws = wb.active
    sheets = wb.sheetnames
    if len(sheets) == 1:
        sheet = sheets[0]
    elif len(sheets) == 2 and 'Version history' in sheets:
        sheet = sheets[0]
    else:
        print(f'Select a sheet from {spreadsheet}:\n')
        actual = [];
        for i,x in enumerate(sheets):
            if x == 'Version history':
                continue
            print(f' {i+1:2}: {x}')
            actual.append(x)

        print('\n  r: return')
        choice = input("\nSelect a sheet > ")
        try:
            if int(choice) > 0 and int(choice) <= len(actual):
                sheet = actual[int(choice)-1]
            else:
                print('No sheet specified')
                return
        except Exception as E:
            print(E)
            print('No sheet specified')
            return
    instrument = str(wb[sheet]['B1'].value).lower()

    if instrument == 'glancing angle':
        print(bold_msg('This is a glancing angle spreadsheet'))
        pinwheel.spreadsheet(spreadsheet, sheet)
    else:
        print(bold_msg('This is a sample wheel spreadsheet'))
        wmb.spreadsheet(spreadsheet, sheet)
    rkvs.set('BMM:automation:type', instrument)







        

run_report('\t'+'areascan')
from BMM.areascan import areascan, as2dat

run_report('\t'+'timescan')
from BMM.timescan import timescan, ts2dat

run_report('\t'+'energystep')
from BMM.energystep import energystep

run_report('\t'+'other plans')
from BMM.plans import tu, td, mvbct, mvrbct, mvbender, mvrbender
from BMM.plans import recover_mirror2, recover_mirror3, recover_mirrors, recover_diagnostics, recover_slits2, recover_slits3

run_report('\t'+'change_mode, change_xtals')
from BMM.modes import change_mode, describe_mode, get_mode, mode, read_mode_data, change_xtals, pds_motors_ready

if os.path.isfile(os.path.join(BMM_CONFIGURATION_LOCATION, 'Modes.json')):
     MODEDATA = read_mode_data()
if BMMuser.pds_mode is None:
     BMMuser.pds_mode = get_mode()


run_report('\t'+'change_edge')
from BMM.edge import show_edges, change_edge
from BMM.functions import approximate_pitch


XDI_record = {'xafs_linx'                        : (True,  'BMM.sample_x_position'),
              'xafs_x'                           : (True,  'BMM.sample_x_position'),
              'xafs_liny'                        : (True,  'BMM.sample_y_position'),
              'xafs_y'                           : (True,  'BMM.sample_y_position'),
              'xafs_lins'                        : (False, 'BMM.sample_s_position'),
              'xafs_linxs'                       : (False, 'BMM.sample_ref_position'),
              'xafs_pitch'                       : (False, 'BMM.sample_pitch_position'),
              'xafs_roll'                        : (False, 'BMM.sample_roll_position'),
              'xafs_roth'                        : (False, 'BMM.sample_roth_position'),
              'xafs_wheel'                       : (False, 'BMM.sample_wheel_position'),
              'xafs_ref'                         : (False, 'BMM.sample_ref_position'),
              'xafs_rots'                        : (False, 'BMM.sample_rots_position'),
              'first_crystal_temperature'        : (False, 'BMM.mono_first_crystal_temperature'),
              'compton_shield_temperature'       : (False, 'BMM.mono_compton_shield_temperature'),
              'dm3_bct'                          : (False, 'BMM.beamline_bct_position'),
              'ring_current'                     : (False, 'BMM.facility_ring_current'),
              'bpm_upstream_x'                   : (False, 'BMM.facility_bpm_upstream_x'),
              'bpm_upstream_y'                   : (False, 'BMM.facility_bpm_upstream_y'),
              'bpm_downstream_x'                 : (False, 'BMM.facility_bpm_downstream_x'),
              'bpm_downstream_y'                 : (False, 'BMM.facility_bpm_downstream_y'),
              'monotc_inboard_temperature'       : (False, 'BMM.mono_tc_inboard'),
              'monotc_upstream_high_temperature' : (False, 'BMM.mono_tc_upstream_high'),
              'monotc_downstream_temperature'    : (False, 'BMM.mono_tc_downstream'),
              'monotc_upstream_low_temperature'  : (False, 'BMM.mono_tc_upstream_low'),
              }
run_report('\t'+'XDI')
from BMM.xdi import write_XDI

run_report('\t'+'machine learning and data evaluation')
from BMM.ml import BMMDataEvaluation
clf = BMMDataEvaluation()

run_report('\t'+'xafs')
from BMM.xafs import howlong, xafs, db2xdi

run_report('\t'+'mono calibration')
from BMM.mono_calibration import calibrate, calibrate_high_end, calibrate_low_end, calibrate_mono

run_report('\t'+'Larch')
from BMM.larch import Pandrosus, Kekropidai
## examples that only work at BMM...
# se = Pandrosus()
# se.fetch('8e293af3-811c-4e96-a4e5-733d0dc77dad', name\='Se metal', mode='transmission')

# seo = Pandrosus()
# seo.fetch('69c35332-6c8a-4f43-9eb2-e5e9cbe7f798', name='SeO2', mode='transmission')

# bunch = Kekropidai(name='Selenium standards')
# bunch.add(se)
# bunch.add(seo)


run_report('\t'+'Demeter')
from BMM.demeter import athena, hephaestus, toprj

run_report('\t'+'telemetry')
from BMM.telemetry import BMMTelementry
tele = BMMTelementry()

run_report('\t'+'user interaction')
# from BMM.wdywtd import WDYWTD
# _do = WDYWTD()
# do = _do.wdywtd
# setup_xrd = _do.do_SetupXRD


from BMM.prompt import MyPrompt, BMM_help, BMM_keys
ip = get_ipython()
ip.prompts = MyPrompt(ip)

if rois.trigger is True:        # set Struck rois from persistent user information
     if len(BMMuser.rois) == 3:
          rois.set(BMMuser.rois)
          rois.select(BMMuser.element)
     rois.trigger = False

if BMMuser.element is None:
     try:
          BMMuser.element = str(rkvs.get('BMM:pds:element'), 'utf-8')
     except:
          pass
     try:
          BMMuser.edge    = str(rkvs.get('BMM:pds:edge'), 'utf-8')
     except:
          pass
if BMMuser.element is not None and with_xspress3 is True: # make sure Xspress3 is configured to measure from the correct ROI
     BMMuser.verify_roi(xs, BMMuser.element, BMMuser.edge)
     #BMMuser.verify_roi(xs1, BMMuser.element, BMMuser.edge)
     show_edges()

xascam._root = os.path.join(BMMuser.folder, 'snapshots')
xrdcam._root = os.path.join(BMMuser.folder, 'snapshots')
anacam._root = os.path.join(BMMuser.folder, 'snapshots')

     
from BMM.edge import all_connected
if all_connected(True) is False:
     print(error_msg('Ophyd connection failure (testing main PDS motors)'))
     print(error_msg('You likely have to restart bsui.'))

     
wmb.folder = BMMuser.folder
wmb.tmpl = os.path.join(os.getenv('HOME'), '.ipython', 'profile_collection', 'startup', 'wheelmacro.tmpl')
pinwheel.folder = BMMuser.folder
pinwheel.tmpl = os.path.join(os.getenv('HOME'), '.ipython', 'profile_collection', 'startup', 'gamacro.tmpl')

RE.msg_hook = BMM_msg_hook

try:
    from bluesky_widgets.utils.streaming import stream_documents_into_runs
    from bluesky_widgets.models.plot_builders import Lines
    from bluesky_widgets.qt.figures import QtFigure
    # model = Lines("xafs_y", ["I0"], max_runs=1)
    # view = QtFigure(model.figure)
    # view.show()
except:
    pass

