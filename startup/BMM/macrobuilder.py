
import os, re, numpy, configparser
from openpyxl import load_workbook
from rich import print as cprint

from BMM.functions      import error_msg, warning_msg, go_msg, url_msg, bold_msg, verbosebold_msg, list_msg, disconnected_msg, info_msg, whisper
from BMM.functions      import isfloat, present_options
from BMM.periodictable  import PERIODIC_TABLE, edge_energy
from BMM.xafs_functions import conventional_grid, sanitize_step_scan_parameters
from BMM.workspace      import rkvs

from BMM import user_ns as user_ns_module
user_ns = vars(user_ns_module)

from BMM.user_ns.base import startup_dir
from BMM.user_ns.bmm  import BMMuser

class BMMMacroBuilder():
    '''A base class for parsing specially constructed spreadsheets and
    generating the corresponding BlueSky plan.

    attributes
    ----------
    basename : str
       basename of the spreadsheet or name of the sheet
    folder : str
       folder containing spreadsheet, usually same as BMMuser.folder
    joiner : str
       string used to construct filenames [-] (_ is a also a good choice)
    source : str
       fully resolved path to spreadsheet
    wb : openpyxl workbook object
       workbook created from spreadsheet
    ws : openpyxl worksheet object
       main sheet of spreadsheet
    measurements : list
       list of disctionaries, one for each row of the spreadsheet
    ini : str
       fully resolved path to INI file
    macro : str
       fully resolved path to plan file
    tab : str
       string used to pythonically format the plan file
    content : str
       accumulated content of plan
    do_first_change : bool
       True is need to begin with a change_edge()
    has_e0_column : bool
       True is this is a very old wheel spreadsheet
    verbose : bool
       True for more comment lines in  the plan 
    totaltime : float
       estimate for the run time of the plan
    deltatime : float
       estimated uncertainty in the total time estimate
    instrument : str
       "sample wheel" or "glancing angle stage"

    Required method
    ---------------
    _write_macro
       generate the text of the BlueSky plan
    get_keywords
       instructions for parsing spreadsheet columns into keywords
    
    '''
    def __init__(self, folder=None):
        self.basename         = None
        self.folder           = None
        self.joiner           = '_'

        self.source           = None
        self.wb               = None
        self.ws               = None
        self.measurements     = list()
        self.ini              = None
        self.macro            = None

        self.tab              = ' ' * 8
        self.content          = ''
        self.do_first_change  = False
        self.has_e0_column    = False
        self.verbose          = False
        self.nreps            = 1

        self.double           = False
        
        self.totaltime        = 0
        self.deltatime        = 0
        self.metadatatime     = 0

        self.description      = ''
        self.ref              = ''
        self.tmpl             = os.path.join(startup_dir, 'tmpl', 'macro.tmpl')
        self.instrument       = None
        self.cleanup          = ''
        self.initialize       = ''

        # parameters that spreadsheets put into .ini files but which are not arguments of the xafs() plan
        self.experiment       = ('default',
                                 'focus', 'measure',
                                 'slot', 'ring',  # ex situ sample wheel
                                 'temperature', 'settle', 'power',  # Linkam, Lakeshore
                                 'spin', 'angle', 'method',  # glancing angle stage
                                 'spinner', 'roi2', 'roi3', 'flat', 'relativep', # resonent reflectivity
                                 'motor1', 'position1', 'motor2', 'position2', 'motor3', 'position3',  # grid automation
                                 'optimize'  # grid, but not in use
        )
        self.flags            = ('snapshots', 'htmlpage', 'usbstick', 'bothways', 'channelcut', 'ththth')
        self.motors           = ('samplex', 'sampley', 'samplep', 'slitwidth', 'slitheight', 'detectorx')
        self.science_metadata = ('url', 'doi', 'cif')

        self.do_opt           = False
        self.optimize         = None
        self.orientation      = 'parallel'
        self.retract          = 10
        self.edgechange       = 'Normal' # 'Quick'
        
    def spreadsheet(self, spreadsheet=None, sheet=None, double=False):
        '''Convert an experiment description spreadsheet to a BlueSky plan.

        Usually called with no arguments, in which case the user will
        be prompted in the bsui shell for a .xlsx file in the data
        directory.  If the .xlsx file has more that one sheet, the
        user will also be prompted for the sheet name.

        Arguments
        =========
        spreadsheet : str
           The fully resolved path to the .xlsx file
        sheet : int/str/None
           The sheet to read from the spreadsheet file. If int, 
           interpret as the index of the sheetnames list, if str 
           interpret as the name from the sheetnames list, if None 
           use the active sheet as determined by openpyxl.

        '''
        if spreadsheet is None:
            spreadsheet = present_options('xlsx')
        if spreadsheet is None:
            error_msg('No spreadsheet specified!')
            return None
        if spreadsheet[-5:] != '.xlsx':
            spreadsheet = spreadsheet+'.xlsx'
        self.folder   = BMMuser.workspace
        self.source   = os.path.join(self.folder, spreadsheet)
        self.basename = os.path.splitext(spreadsheet)[0]
        self.wb       = load_workbook(self.source, data_only=True, read_only=True);
        self.measurements = list()
        # self.do_first_change = False
        # self.close_shutters  = True

        self.double = double

        #print(f'-----{sheet}')
        if sheet is None:
            self.ws = self.wb.active
        elif type(sheet) is int:
            this = self.wb.sheetnames[sheet-1]
            self.ws = self.wb[this]
            self.basename = this
        elif sheet in self.wb.sheetnames:
            self.ws = self.wb[sheet]
            self.basename = sheet
        else:
            self.ws = self.wb.active
        self.basename = re.sub('[ \-+*/]+', '_', self.basename)
        self.ini      = os.path.join(self.folder, self.basename+'.ini')
        self.macro    = os.path.join(self.folder, self.basename+'_macro.py')

        ## this is looking for presence/absence of a column labeled "optimize", which is experimental on 13 January 2023
        #self.do_opt = False
        #if 'wheel' in str(type(self)) and 'Optimize' in self.ws['U5'].value:
        #    self.do_opt = True

        ## this is trying to deal with very early spreadsheets which had an e0 column which no longer exists 
        if self.ws['H5'].value.lower() == 'e0': 
            self.has_e0_column = True

        ## this is dealing with single and double ring sample wheels
        if double is True:
            #self.do_first_change = self.truefalse(self.ws['H2'].value, 'firstchange')
            self.close_shutters  = self.truefalse(self.ws['K2'].value, 'closeshutters')
            self.append_element  = str(self.ws['M2'].value)
            self.nreps           = self.ws['O2'].value
        else:
            #self.do_first_change = self.truefalse(self.ws['G2'].value, 'firstchange')
            self.close_shutters  = self.truefalse(self.ws['J2'].value, 'closeshutters')
            self.append_element  = str(self.ws['L2'].value)
            self.nreps           = self.ws['N2'].value
            if str(self.ws['G3'].value) == 'perp':
                self.orientation = 'perpendicular' # only GA
            else:
                self.orientation = 'parallel' # only GA                
            if self.ws['N3'].value is not None:
                if 'Retraction'  in str(self.ws['M3'].value):
                    self.retract = abs(self.ws['N3'].value)      # only GA
                if 'Edge change' in str(self.ws['M3'].value):
                    self.edgechange = str(self.ws['N3'].value)      # only GA
        if self.nreps is None:
            self.nreps = 1
        else:
            self.nreps = int(self.nreps)

        self.instrument = str(self.ws['B1'].value).lower()
        self.version = str(self.ws['B2'].value).lower()

        isok, explanation, reference = self.read_spreadsheet()
        if isok is False:
            print('\n' + error_msg(explanation))
            print(f'See: {reference}')
            return None
        self.write_macro()
        return 0

    def truefalse(self, value, keyword):
        '''Interpret certain strings from the spreadsheet as True/False'''
        if value is None:
            if keyword.lower() in ('bothways', 'ththth'):
                return False
            else:
                return True  # self.measurements[0]['measure']
        if type(value) is bool:
            return value
        elif str(value).lower() == '=true()':
            return True
        elif str(value).lower() == 'true':
            return True
        elif str(value).lower() == 'yes':
            return True
        elif type(value) is int and value == 1:
            return True
        else:
            return False

    def nonezero(self, value):
        '''Interpret None as being 0 valued'''
        if value is None:
            return(0)
        elif type(value) == str:
            return(0)
        else:
            return(float(value))

    def zeronone(self, value):
        '''Interpret None as being None valued'''
        if value is None:
            return(None)
        elif type(value) == str:
            return(None)
        else:
            return(float(value))

    def escape_quotes(self, value):
        if value is None:
            return ''
        if type(value) is str and value == 'None':
            return ''
        value = value.replace('\\', "\\\\").replace('\'', "\\'").replace('"', '\\"')
        return value

    def is_empty(self, value):
        if value is None:
            return True
        if type(value) is str and value.strip() == '':
            return True
        return False

    def check_spinner(self, value):
        if type(value)is not int:
            error_msg(f"Spinner number must be an integer between 1 and 8. You said {value}")
            return False
        if value <1 or value > 8:
            error_msg(f"Spinner number must be an integer between 1 and 8. You said {value}")
            return False
        return True 
    
    def check_limit(self, motor, value):
        '''Perform a sanity check on a requested motor position.
        Return False if there is a problem.
        '''
        if type(motor) is str:
            if motor in user_ns:
                motor = user_ns[motor]
            else:
                error_msg(f'"{motor}" is not a valid motor name.')
                return False
        if value > motor.limits[1]:
            error_msg(f"A requested {motor.name} position ({value}) is greater than the high limit ({motor.limits[1]})")
            return False
        if value < motor.limits[0]:
            error_msg(f"A requested {motor.name} position ({value}) is less than the low limit ({motor.limits[0]})")
            return False
        return True 

    def check_temp(self, stage, value):
        '''Perform a sanity check on a requested heating/cooling stage temperature.
        Return False if there is a problem.
        '''
        name, units = stage.name.capitalize(), 'C'
        if 'lakeshore' in stage.name.lower():
            name, units = 'Displex', 'K'
        if value > stage.limits[1]:
            error_msg(f"A requested {name} temperature ({value}{units}) is greater than the high limit ({stage.limits[1]}{units})")
            return False
        if value < stage.limits[0]:
            error_msg(f"A requested {name} temperature ({value}{units}) is less than the low limit ({stage.limits[0]}{units})")
            return False
        return True

    
        
    def ini_sanity(self, default):
        '''Sanity checks for the default line from the spreadsheet.

        1. experimenters is a string (BMMuser.name)
        2. sample, prep, and comment are not empty strings (set to '...')
        3. nscans is an integer (set to 1)
        4. start is an integer or "next"
        5. mode is string (set to 'transmission')
        6. element is an element (bail)
        7. edge is k, l1, l2, or l3 (bail)

        To do:
          * booleans are interpretable as booleans
          * focused is focused or unfocused
          * bounds, steps, times are sensible
          * x, y, slits are floats and sensible for the respective ranges of motion

        '''

        message = ''
        unrecoverable = False

        if 'mode' not in default:
            if 'glancing angle' in self.instrument:
                default['mode'] = 'xs'
            else:
                default['mode'] = 'transmission'

        if default['filename'] is None or str(default['filename']).strip() == '':
            default['filename'] = 'filename'

        #if default['experimenters'] is None or str(default['experimenters']).strip() == '':
        default['experimenters'] = BMMuser.experimenters

        defaultdefaults = {'bounds': '-200  -30  -10 15.5  570', 'steps': '10  2  0.25  0.05k', 'times': '1 1 1 1'}
        for k in ('bounds', 'steps', 'times'):
            if default[k] is None or str(default[k]).strip() == '':
                default[k] = defaultdefaults[k]

        for k in ('sample', 'prep', 'comment'):
            if default[k] is None or str(default[k]).strip() == '':
                default[k] = '...'
            if '%' in default[k]:
                default[k] = default[k].replace('%', '%%')

        try:
            default['nscans'] = int(default['nscans'])
        except:
            default['nscans'] = 1

        try:
            default['start'] = int(default['start'])
        except:
            default['start'] = 'next'

        # if default['mode'] is None or str(default['mode']).strip() == '':
        #    default['mode'] = 'transmission'

        el = str(default['element']).capitalize()
        ed = str(default['edge']).lower()
        if el not in re.split('\s+', PERIODIC_TABLE): # see 06-periodic table 
            message += '\nDefault entry for element is not recognized.'
            unrecoverable = True

        if ed not in ('k', 'l1', 'l2', 'l3'):
            message += '\nDefault entry for edge is not recognized.'
            unrecoverable = True

        dcm = user_ns['dcm']
        ee = edge_energy(el, ed)
        if ee is not None:
            if dcm._crystal == '111' and ee > 21200:
                message += f'\nCannot measure {el} {ed} edge on the {dcm._crystal} crystals.'
                unrecoverable = True
            if dcm._crystal == '311' and ee < 5500:
                message += f'\nCannot measure {el} {ed} edge on the {dcm._crystal} crystals.'
                unrecoverable = True
            
            
        # try:
        #     default['e0'] = float(default['e0'])
        # except:
        #     default['e0'] = edge_energy(default['element'], default['edge'])

        if unrecoverable:
            error_msg(message)
            default = None
        return default

    def read_spreadsheet(self):
        '''Slurp up the content of the spreadsheet and write the default control file
        '''
        print('Reading spreadsheet: %s' % self.source)
        count = 0
        isok, explanation = True, ''

        for row in self.ws.rows:
            count += 1
            if count < 6:
                continue
            defaultline = False
            if count == 6:
                defaultline = True
            if count > 300:
                break

            this = self.get_keywords(row, defaultline)
            if self.skip_row(this) and not defaultline:
                continue
            self.measurements.append(this)

            # check that scan parameters make sense
            bounds = self.measurements[-1]['bounds']
            if type(bounds) is str and bounds != 'None' and bounds.strip() != '':
                b = re.split('[ ,]+', self.measurements[-1]['bounds'].strip())
            else:
                b = re.split('[ ,]+', self.measurements[0]['bounds'].strip())

            steps = self.measurements[-1]['steps']
            if type(steps) is str and steps != 'None' and steps.strip() != '':
                s = re.split('[ ,]+', self.measurements[-1]['steps'].strip())
            else:
                s = re.split('[ ,]+', self.measurements[0]['steps'].strip())

            times = self.measurements[-1]['times']
            if type(times) is str and times != 'None' and times.strip() != '':
                t = re.split('[ ,]+', self.measurements[-1]['times'].strip())
            else:
                t = re.split('[ ,]+', self.measurements[0]['times'].strip())
                
            (problem, text, reference) = sanitize_step_scan_parameters(b, s, t)
            if problem is True:
                isok = False
                explanation += bold_msg(f'\nrow {count}, sample {self.measurements[-1]["filename"]}:\n') + text

            def float_or_string(l):
                out = []
                for thing in l:
                    try:
                        out.append(float(thing))
                    except:
                        out.append(thing)
                return out
                        
            el = self.measurements[-1]['element'] or self.measurements[0]['element']
            ed = self.measurements[-1]['edge']    or self.measurements[0]['edge']
            e0 = edge_energy(el, ed)
            (grid, inttime, time, delta) = conventional_grid(bounds=float_or_string(b), steps=float_or_string(s), times=float_or_string(t), element=el, edge=ed) #, e0=e0)
            if any(it > 20 for it in inttime):
                isok = False
                text = error_msg('\tYour scan asks for an integration time greater than 20 seconds, which the ion chamber electrometer cannot accommodate.')
                explanation += bold_msg(f'\nrow {count}, sample {self.measurements[-1]["filename"]}:\n') + text
                
            
                
        self.calls_to_xafs = 0
        for m in self.measurements:
            if m['default'] is False and  self.skip_row(m) is False:
                self.calls_to_xafs += 1
        self.calls_to_xafs = self.calls_to_xafs * self.nreps
        return(isok, explanation, reference)
        # pp.pprint(self.measurements)


    def skip_row(self, m):
        #####################################################
        # all the reasons to skip a line in the spreadsheet #
        #####################################################
        if 'slot' in m and type(m['slot']) is not int:
            return True
        #if 'temperature' in m and type(m['temperature']) is not float:
        #    return True
        #print('>', m['filename'], '<')
        #print(type(m['filename']))
        #print(self.is_empty(m['filename']))
        #print()
        if self.is_empty(m['filename']):
            return True
        if m['filename'] == 'None':  # where does an empty cell get turned into "None"??
            return True
        if  self.truefalse(m['measure'], 'measure') is False:
            return True
        if m['nscans'] is not None and m['nscans'] < 1:
            return True
        return False

    def skip_keyword(self, k):
        '''Identify all the keywords that should NOT be captured in the xafs() call.'''
        if k in self.experiment or k in self.flags or k in self.motors or k in self.science_metadata:
            return True
        return False

    def make_filename(self, m):
        '''Construct a filename with element and edge symbols, if required.

        Also remove troublesome characters from the filename.
             * / \ ? % : | " < >

        Simply removing those characters makes many things simpler,
        including stuff from os.path and any shell interactions.  Also
        makes it possible to put data files onto a DOS formatted USB
        stick (for the old skool cool).

        '''
        fname = m['filename']
        #bad_chars = ('*', '/', '\\', '?', '%', ':', '|', '"', '<', '>')
        fname = re.sub('[*/<>%?:"|\\\\]', self.joiner, fname)
        el = self.measurements[0]['element']
        ed = self.measurements[0]['edge']
        t = ''
        if 'temperature' in self.measurements[0]:
            t  = f'{int(self.measurements[0]["temperature"]):03d}'
        if 'element' in m:
            el = m['element']
        if 'edge' in m:
            ed = m['edge']
        if 'temperature' in m:
            t = str(int(m['temperature']))
        if self.append_element.lower() == 'element at beginning':
            fname = el + self.joiner + fname
        elif self.append_element.lower() == 'element at end':
            fname = fname + self.joiner + el
        elif self.append_element.lower() == 'element+edge at beginning':
            fname = el + self.joiner + ed + self.joiner + fname
        elif self.append_element.lower() == 'element+edge at end':
            fname = fname + self.joiner + el + self.joiner + ed
        elif self.append_element.lower() == 'temperature at beginning':
            fname = t + self.joiner + fname
        elif self.append_element.lower() == 'temperature at end':
            fname = fname + self.joiner + t
        elif self.append_element.lower() == 'temperature+element at beginning':
            fname = el + self.joiner + t + self.joiner + fname
        elif self.append_element.lower() == 'temperature+element at end':
            fname = fname + self.joiner + el + self.joiner + t
        elif self.append_element.lower() == 'temperature+element+edge at beginning':
            fname = el + self.joiner + ed + self.joiner + t + self.joiner + fname
        elif self.append_element.lower() == 'temperature+element+edge at end':
            fname = fname + self.joiner + el + self.joiner + ed + self.joiner + t

        if 'relativep' in m:
            fname = fname + self.joiner + str(m['relativep']) + 'deg'
            
        return fname

    def do_change_edge(self, el, ed, focus, tab):
        if self.edgechange == 'Quick':
            text = f"{tab}yield from quick_change('{el}', edge='{ed}', focus={focus})\n"
        else:
            text = f"{tab}yield from change_edge('{el}', edge='{ed}', focus={focus})\n"
        time = 5.0
        inrange = True
        ee = edge_energy(el, ed)
        if ee > 23500 or ee < 3500:
            error_msg(f'\nThe {el} {ed} energy {ee:.1f} is outside the available range at BMM.')
            print('You probably have the edge set incorrectly in your spreadsheet.\n')
            inrange = False
        return(text, time, inrange)

    def check_edge(self):
        first_element, first_edge = self.measurements[1]['element'], self.measurements[1]['edge']
        first_focus = self.measurements[1]['focus']
        if first_element is None:
            first_element = self.measurements[0]['element']
        if first_edge is None:
            first_edge = self.measurements[0]['edge']
        if first_focus is None:
            first_focus = self.measurements[0]['focus']
        if first_focus == 'focused':
            first_focus = 'True'
        else:
            first_focus = 'False'
        #current_element = rkvs.get('BMM:user:element').decode('utf-8')
        #current_edge    = rkvs.get('BMM:user:edge').decode('utf-8')

        text  = self.tab + "## change edge before starting, if needed...\n"
        text += self.tab + f"if not_at_edge('{first_element}', '{first_edge}'):\n"
        if self.edgechange == 'Quick':
            text += self.tab + f"    yield from quick_change('{first_element}', edge='{first_edge}', focus={first_focus})\n\n"
        else:
            text += self.tab + f"    yield from change_edge('{first_element}', edge='{first_edge}', focus={first_focus})\n\n"

        return text

    
    def estimate_time(self, m, el, ed):
        '''Approximate the time contribution from the current row'''
        if type(m['bounds']) is str and m['bounds'].strip() != '':
            b = re.split('[ ,]+', m['bounds'].strip())
        else:
            b = re.split('[ ,]+', self.measurements[0]['bounds'].strip())
        if type(m['steps']) is str and m['steps'].strip() != '':
            s = re.split('[ ,]+', m['steps'].strip())
        else:
            s = re.split('[ ,]+', self.measurements[0]['steps'].strip())
        if type(m['times']) is str and m['times'].strip() != '':
            t = re.split('[ ,]+', m['times'].strip())
        else:
            t = re.split('[ ,]+', self.measurements[0]['times'].strip())

        b = [float(x) if isfloat(x) else x for x in b]
        s = [float(x) if isfloat(x) else x for x in s]
        t = [float(x) if isfloat(x) else x for x in t]

        (e, t, at, delta) = conventional_grid(bounds=b, steps=s, times=t, e0=edge_energy(el, ed), element=el, edge=ed, ththth=False)

        if type(m['nscans']) is int:
            nsc = m['nscans']
        else:
            nsc = self.measurements[0]['nscans']
        self.totaltime += at * nsc * self.nreps
        self.deltatime += delta*delta * self.nreps
        tele = user_ns['tele']
        try:
            self.metadatatime += tele.value(el, 'visual')
        except:
            pass
        try:
            if m['mode'] in ('fluorescence', 'flourescence', 'both', 'xs', 'xs1'):
                self.metadatatime += tele.value(el, 'xrf')
        except:
            pass
        #print(at, nsc, self.nreps, tele.value(el, 'visual'), tele.value(el, 'xrf'))
        
    def write_ini_and_plan(self):
        #################################
        # write out the master INI file #
        #################################
        config = configparser.ConfigParser()
        default = self.measurements[0].copy()
        #          things in the spreadsheet but not in the INI file
        # for k in ('default', 'slot', 'measure', 'spin', 'focus', 'method',
        #           'samplep', 'samplex', 'sampley', 'slitwidth', 'slitheight', 'detectorx',
        #           'settle', 'power', 'temperature', 'motor1', 'position1', 'motor2', 'position2', 'motor3', 'position3',
        #           'optimize', 'roi2', 'roi3', 'spinner', 'flat', 'relativep'):
        for k in self.experiment + self.motors + self.flags + self.science_metadata:
            default.pop(k, None)
        default['url'] = '...'
        default['doi'] = '...'
        default['cif'] = '...'
        # if 'double' in self.instrument.lower():
        #     default['experimenters'] = self.ws['F1'].value  # top line of xlsx file
        # else:
        #     default['experimenters'] = self.ws['E1'].value  # top line of xlsx file
        default = self.ini_sanity(default)
        if default is None:
            error_msg(f'Could not interpret {self.source} as a wheel macro.')
            return
        #import pprint
        #pprint.pprint(default)
        config.read_dict({'scan': default})

        ## write ini file and macro script to workspace
        ## both will be copied to storage later
        ## see xlsx() in user_ns/bmm_end.py
        with open(self.ini, 'w') as configfile:
            config.write(configfile)
        whisper('Wrote default INI file: %s' % self.ini)

        #########################################################
        # write the full macro to a file and %run -i that file  #
        #########################################################
        with open(self.tmpl) as f:
            text = f.readlines()
        fullmacro = ''.join(text).format(folder      = self.folder,
                                         base        = self.basename,
                                         content     = self.content,
                                         description = self.description,
                                         instrument  = self.instrument,
                                         cleanup     = self.cleanup,
                                         initialize  = self.initialize)
        o = open(self.macro, 'w')
        o.write(fullmacro)
        o.close()
        ## I think this will never be called by queueserver
        from IPython import get_ipython
        ipython = get_ipython()
        ipython.run_line_magic('run',  f' -i \'{self.macro}\'')
        whisper('Wrote and read macro file: %s' % self.macro)

    def finish_macro(self):
        #######################################
        # explain to the user what to do next #
        #######################################
        cprint(f'\nYour new {BMMuser.instrument} plan is called:  [yellow2]{self.basename}_macro[/yellow2]')
        cprint(f'\nVerify:  [yellow2]{self.basename}_macro??[/yellow2]')
        if 'glancing angle' in self.instrument:
            cprint(f'Run:     [yellow2]RE({self.basename}_macro())[/yellow2]')
            cprint(f'Add ref: [yellow2]RE({self.basename}_macro(ref=True))[/yellow2]')
        else:
            cprint(f'Run:     [yellow2]RE({self.basename}_macro())[/yellow2]')
            #cprint('Dryrun:  [yellow2]RE({self.basename}_macro(dryrun=True)[/yellow2]')

        alltime = self.totaltime + self.metadatatime/60
        hours = int(alltime/60)
        minutes = int(alltime - hours*60)
        self.deltatime = numpy.sqrt(self.deltatime)
        print(f'\nApproximate time: {hours} hours, {minutes} minutes +/- {self.deltatime:.1f} minutes')

    def write_macro(self):
        '''Write INI file and a BlueSky plan from a spreadsheet.

        Call the subclass' _write_macro to generate the text of the plan.

        '''
        self.totaltime, self.deltatime, self.metadatatime = 0, 0, 0
        self.content = ''
        success = self._write_macro()     # populate self.content
        if success is False: return
        # write_ini_and_plan uses self.measurements and self.content
        self.write_ini_and_plan()
        self.finish_macro()
