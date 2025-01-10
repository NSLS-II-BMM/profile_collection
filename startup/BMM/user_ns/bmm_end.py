try:
    from bluesky_queueserver import is_re_worker_active
except ImportError:
    # TODO: delete this when 'bluesky_queueserver' is distributed as part of collection environment
    def is_re_worker_active():
        return False

import os, textwrap
from BMM.functions import run_report, disconnected_msg, error_msg, whisper, boxedtext, verbosebold_msg, proposal_base
from BMM.workspace import rkvs

from BMM import user_ns as user_ns_module
user_ns = vars(user_ns_module)

from BMM.user_ns.bmm import BMM_CONFIGURATION_LOCATION, BMMuser
from BMM.user_ns.motors import mcs8_motors, xafs_motors

run_report(__file__, text='import the rest of the things')

run_report('\t'+'resting state')
from BMM.resting_state import resting_state, resting_state_plan, end_of_macro

run_report('\t'+'motor status reporting')
from BMM.motor_status import motor_status, ms # , motor_metadata, xrd_motors, xrdm

run_report('\t'+'FMBO motor tools')
from BMM.fmbo import FMBO_status

from BMM.user_ns.base import profile_configuration
from BMM.desc_string  import set_desc_strings
if profile_configuration.getboolean('miscellaneous', 'set_desc_strings'):
    run_report('\t'+'setting motor description strings')
    set_desc_strings()

run_report('\t'+'suspenders')
from BMM.suspenders import BMM_suspenders, BMM_clear_to_start, BMM_clear_suspenders

run_report('\t'+'linescan, rocking curve, slit_height, find_slot, pluck')
from BMM.linescans import linescan, pluck, rocking_curve, slit_height, ls2dat, find_slot, rectangle_scan

run_report('\t'+'kafka')
from BMM.kafka import close_line_plots, close_plots, kafka_message, preserve

run_report('\t'+'support for wafer samples')
from BMM.wafer import Wafer
wafer = Wafer()



###########################################################################################
#  _____  _       ___   _   _ _____ _____ _   _ _____    ___   _   _ _____  _      _____  #
# |  __ \| |     / _ \ | \ | /  __ \_   _| \ | |  __ \  / _ \ | \ | |  __ \| |    |  ___| #
# | |  \/| |    / /_\ \|  \| | /  \/ | | |  \| | |  \/ / /_\ \|  \| | |  \/| |    | |__   #
# | | __ | |    |  _  || . ` | |     | | | . ` | | __  |  _  || . ` | | __ | |    |  __|  #
# | |_\ \| |____| | | || |\  | \__/\_| |_| |\  | |_\ \ | | | || |\  | |_\ \| |____| |___  #
#  \____/\_____/\_| |_/\_| \_/\____/\___/\_| \_/\____/ \_| |_/\_| \_/\____/\_____/\____/  #
###########################################################################################


run_report('\tglancing angle stage')
from BMM.glancing_angle import GlancingAngle, GlancingAngleMacroBuilder
ga = GlancingAngle('XF:06BMB-CT{DIODE-Local:1}', name='glancing angle stage')

gawheel = GlancingAngleMacroBuilder()
gawheel.description = 'the glancing angle stage'
gawheel.instrument  = 'glancing angle'
gawheel.folder = BMMuser.workspace
gawheel.cleanup = 'yield from mv(xafs_x, samx, xafs_y, samy, xafs_pitch, samp, xafs_det, 205)\n        yield from ga.reset()'
gawheel.initialize = '''samx, samy, samp = xafs_x.position, xafs_y.position, xafs_pitch.position
    if ga.ready_to_start() is not True:
        return(yield from null())'''


#########################################################################################
# ___  ___  ___  _____ ______ _____  ______ _   _ _____ _    ______ ___________  _____  #
# |  \/  | / _ \/  __ \| ___ \  _  | | ___ \ | | |_   _| |   |  _  \  ___| ___ \/  ___| #
# | .  . |/ /_\ \ /  \/| |_/ / | | | | |_/ / | | | | | | |   | | | | |__ | |_/ /\ `--.  #
# | |\/| ||  _  | |    |    /| | | | | ___ \ | | | | | | |   | | | |  __||    /  `--. \ #
# | |  | || | | | \__/\| |\ \\ \_/ / | |_/ / |_| |_| |_| |___| |/ /| |___| |\ \ /\__/ / #
# \_|  |_/\_| |_/\____/\_| \_|\___/  \____/ \___/ \___/\_____/___/ \____/\_| \_|\____/  #
#########################################################################################

run_report('\t\tspreadsheet/instrument support functions')
from BMM.functions import present_options, bold_msg
from openpyxl import load_workbook

def xlsx():
    '''Prompt for a macro building spreadsheet for any instrument. Use the
    content of cell B1 to direct this spreadsheet to the correct builder.

    if cell B1 is "Glancing angle" --> build a glancing angle macro

    if cell B1 is "Linkam" --> build a Linkam stage macro

    if cell B1 is "Lakeshore" --> build a Lakeshore + Displex macro

    if cell B1 is "Grid" --> build a generic grid macro

    if cell B1 is "Sample wheel" --> build a sample wheel macro

    if cell B1 is empty --> build a sample wheel macro

    Then prompt for the sheet, if there is more than 1 sheet in the
    spreadsheet file.  Finally prompt for the tab, if there is more
    than 1 tab in the selected spreadsheet.

    '''
    spreadsheet = present_options('xlsx')
    if spreadsheet is None:
        print(error_msg('No spreadsheet specified!'))
        return None

    wb = load_workbook(os.path.join(BMMuser.workspace, spreadsheet), read_only=True, data_only=True)


    version    = int(str(wb[wb.sheetnames[0]]['B2'].value).split(" ")[1])
    if version < 13:
        text = f'''\nHey! That spreadsheet is too old! (found v{version}, must be v13 or greater)

Recreate {spreadsheet} using an example from
the templates folder of your data directory.
'''
        boxedtext(f'{spreadsheet} cannot be imported', text, 'red', width=77)
        return

    #ws = wb.active
    sheets = wb.sheetnames
    if len(sheets) == 1:
        sheet = sheets[0]
    elif len(sheets) == 2 and 'Version history' in sheets:
        sheet = sheets[0]
    else:
        print(f'Select a sheet from {spreadsheet}:\n')
        actual = [];
        for i,x in enumerate(sheets):
            if x == 'Version history':
                continue
            print(f' {i+1:2}: {x}')
            actual.append(x)

        print('\n  r: return')
        choice = input("\nSelect a sheet > ")
        try:
            if int(choice) > 0 and int(choice) <= len(actual):
                sheet = actual[int(choice)-1]
            else:
                print('No sheet specified')
                return
        except Exception as E:
            print(E)
            print('No sheet specified')
            return
    instrument = str(wb[sheet]['B1'].value).lower()

    if instrument.lower() == 'glancing angle':
        print(bold_msg('This is a glancing angle spreadsheet'))
        gawheel.spreadsheet(spreadsheet, sheet)
        BMMuser.instrument = 'glancing angle stage'
    elif instrument.lower() == 'double wheel':
        print(bold_msg('This is a double sample wheel spreadsheet'))
        wmb.spreadsheet(spreadsheet, sheet, double=True)
        BMMuser.instrument = 'double wheel'
    elif instrument.lower() == 'linkam':
        print(bold_msg('This is a Linkam spreadsheet'))
        lmb.spreadsheet(spreadsheet, sheet)
        BMMuser.instrument = 'Linkam stage'
    elif instrument.lower() == 'lakeshore':
        print(bold_msg('This is a LakeShore spreadsheet'))
        lsmb.spreadsheet(spreadsheet, sheet)
        BMMuser.instrument = 'LakeShore 331'
    elif instrument.lower() == 'grid':
        print(bold_msg('This is a motor grid spreadsheet'))
        gmb.spreadsheet(spreadsheet, sheet)
        BMMuser.instrument = 'motor grid'
    elif instrument.lower() == 'sample wheel':
        print(bold_msg('This is a single wheel spreadsheet'))
        print(error_msg('Single wheel spreadsheets have been retired.'))
        print(error_msg('Use a double wheel spreadsheet, instead.'))
    else:
        print(bold_msg('This is a double sample wheel spreadsheet'))
        wmb.spreadsheet(spreadsheet, sheet, double=True)
        BMMuser.instrument = 'double wheel'

    rkvs.set('BMM:automation:type', instrument)
    kafka_message({'copy': True,
                   'file': os.path.join(BMMuser.workspace, spreadsheet),
                   'target': proposal_base(), })
    kafka_message({'copy': True,
                   'file': os.path.join(BMMuser.workspace, f'{sheet}.ini'),
                   'target': proposal_base(), })
    kafka_message({'copy': True,
                   'file': os.path.join(BMMuser.workspace, f'{sheet}_macro.py'),
                   'target': proposal_base(), })


def set_instrument(instrument=None):
    if instrument is None:
        print('''
  1: Sample wheel
  2: Glancing angle stage
  3: Linkam stage
  4: Displex and LakeShore
  5: Motor grid

  r: return
''')
        actual = ['', 'sample wheel', 'glancing angle', 'linkam', 'lakeshore', 'grid']
        choice = input("\nSelect an instrument > ")
        try:
            if int(choice) > 0 and int(choice) <= 6:
                instrument = actual[int(choice)]
            else:
                instrument = 'double wheel'
        except:
            instrument = 'double wheel'

    if instrument.lower() == 'glancing angle':
        print(bold_msg('Setting instrument as glancing angle stage'))
        BMMuser.instrument = 'glancing angle stage'
    elif instrument.lower() == 'sample wheel':
        print(bold_msg('Setting instrument as double sample wheel'))
        BMMuser.instrument = 'double wheel'
    elif instrument.lower() == 'linkam':
        print(bold_msg('Setting instrument as Linkam stage'))
        BMMuser.instrument = 'Linkam stage'
    elif instrument.lower() == 'lakeshore':
        print(bold_msg('Setting instrument as LakeShore 331'))
        BMMuser.instrument = 'LakeShore'
    elif instrument.lower() == 'grid':
        print(bold_msg('Setting instrument as sample grid'))
        BMMuser.instrument = 'motor grid'
    else:
        print(bold_msg('Default instrument choice: sample wheel'))
        BMMuser.instrument = 'double wheel'
    rkvs.set('BMM:automation:type', BMMuser.instrument)
    
    

###########################################################################################
# ______ _   _ _____ _____ _____ _   _  ______ _____ _     _____ _   _ _____________   __ #
# | ___ \ | | |  _  |_   _|  _  | \ | | |  _  \  ___| |   |_   _| | | |  ___| ___ \ \ / / #
# | |_/ / |_| | | | | | | | | | |  \| | | | | | |__ | |     | | | | | | |__ | |_/ /\ V /  #
# |  __/|  _  | | | | | | | | | | . ` | | | | |  __|| |     | | | | | |  __||    /  \ /   #
# | |   | | | \ \_/ / | | \ \_/ / |\  | | |/ /| |___| |_____| |_\ \_/ / |___| |\ \  | |   #
# \_|   \_| |_/\___/  \_/  \___/\_| \_/ |___/ \____/\_____/\___/ \___/\____/\_| \_| \_/   #
###########################################################################################
                                                                                      

run_report('\t'+'other plans')
from BMM.plans import tu, td, mvbct, mvrbct, mvbender, mvrbender, move, mover
from BMM.plans import recover_mirror2, recover_mirror3, recover_mirrors, recover_diagnostics, recover_slits2, recover_slits3

run_report('\t'+'change_mode, change_xtals, PDS utilities')
from BMM.modes import change_mode, describe_mode, get_mode, mode, read_mode_data, MODEDATA, change_xtals, pds_motors_ready, table_height

if os.path.isfile(os.path.join(BMM_CONFIGURATION_LOCATION, 'Modes.json')):
     MODEDATA = read_mode_data()
if BMMuser.pds_mode is None:
     BMMuser.pds_mode = get_mode()

run_report('\t'+'change_edge')
from BMM.edge import show_edges, change_edge, xrd_mode
from BMM.functions import approximate_pitch

run_report('\t'+'mono calibration')
from BMM.mono_calibration import calibrate, calibrate_high_end, calibrate_low_end, calibrate_mono, calibrate_pitch


###############################################################
#  _____ _____   ___   _   _   _______   _______ _____ _____  #
# /  ___/  __ \ / _ \ | \ | | |_   _\ \ / / ___ \  ___/  ___| #
# \ `--.| /  \// /_\ \|  \| |   | |  \ V /| |_/ / |__ \ `--.  #
#  `--. \ |    |  _  || . ` |   | |   \ / |  __/|  __| `--. \ #
# /\__/ / \__/\| | | || |\  |   | |   | | | |   | |___/\__/ / #
# \____/ \____/\_| |_/\_| \_/   \_/   \_/ \_|   \____/\____/  #
###############################################################


XDI_record = {'xafs_linx'                        : (True,  'Sample.x'),
              'xafs_x'                           : (True,  'Sample.x'),
              'xafs_liny'                        : (True,  'Sample.y'),
              'xafs_y'                           : (True,  'Sample.y'),
              'xafs_lins'                        : (True,  'Sample.SDD_position'),
              'xafs_det'                         : (True,  'Sample.SDD_position'),
              'xafs_linxs'                       : (False, 'BMM.sample_ref_position'),
              'xafs_pitch'                       : (False, 'Sample.pitch'),
              'xafs_roll'                        : (False, 'BMM.sample_roll_position'),
              'xafs_wheel'                       : (False, 'BMM.sample_wheel_position'),
              'xafs_ref'                         : (False, 'BMM.sample_ref_position'),
              'xafs_rots'                        : (False, 'BMM.sample_rots_position'),
              'first_crystal_temperature'        : (False, 'BMM.mono_first_crystal_temperature'),
              'compton_shield_temperature'       : (False, 'BMM.mono_compton_shield_temperature'),
              'dm3_bct'                          : (False, 'BMM.beamline_bct_position'),
              'ring_current'                     : (False, 'BMM.facility_ring_current'),
              'bpm_upstream_x'                   : (False, 'BMM.facility_bpm_upstream_x'),
              'bpm_upstream_y'                   : (False, 'BMM.facility_bpm_upstream_y'),
              'bpm_downstream_x'                 : (False, 'BMM.facility_bpm_downstream_x'),
              'bpm_downstream_y'                 : (False, 'BMM.facility_bpm_downstream_y'),
              'monotc_inboard_temperature'       : (False, 'BMM.mono_tc_inboard'),
              'monotc_upstream_high_temperature' : (False, 'BMM.mono_tc_upstream_high'),
              'monotc_downstream_temperature'    : (False, 'BMM.mono_tc_downstream'),
              'monotc_upstream_low_temperature'  : (False, 'BMM.mono_tc_upstream_low'),
              }

# suppress some uninteresting messages from lib/python3.7/site-packages/hdf5plugin/__init__.py
# import logging
# logging.getLogger("hdf5plugin").setLevel(logging.ERROR) # no longer needed, I guess...
run_report('\t'+'xafs')
from BMM.xafs import howlong, xafs, xanes
from BMM.xafs_functions import xrfat
from BMM.dossier import lims

run_report('\t'+'areascan')
from BMM.areascan import areascan, as2dat, fetch_areaplot

run_report('\t'+'timescan')
from BMM.timescan import timescan, ts2dat, sead

# run_report('\t'+'energystep')
# from BMM.energystep import energystep

run_report('\t'+'raster scans')
from BMM.raster import raster #, difference_data


################################################################################################################
#  _______   ________ ___________ ________  ___ _____ _   _ _____   _____ _   _____________ ___________ _____  #
# |  ___\ \ / /| ___ \  ___| ___ \_   _|  \/  ||  ___| \ | |_   _| /  ___| | | | ___ \ ___ \  _  | ___ \_   _| #
# | |__  \ V / | |_/ / |__ | |_/ / | | | .  . || |__ |  \| | | |   \ `--.| | | | |_/ / |_/ / | | | |_/ / | |   #
# |  __| /   \ |  __/|  __||    /  | | | |\/| ||  __|| . ` | | |    `--. \ | | |  __/|  __/| | | |    /  | |   #
# | |___/ /^\ \| |   | |___| |\ \ _| |_| |  | || |___| |\  | | |   /\__/ / |_| | |   | |   \ \_/ / |\ \  | |   #
# \____/\/   \/\_|   \____/\_| \_|\___/\_|  |_/\____/\_| \_/ \_/   \____/ \___/\_|   \_|    \___/\_| \_| \_/   #
################################################################################################################
                                                                                                           


run_report('\t'+'Larch')
from BMM.larch_interface import Pandrosus, Kekropidai
## examples that only work at BMM...
# se = Pandrosus()
# se.fetch('8e293af3-811c-4e96-a4e5-733d0dc77dad', name\='Se metal', mode='transmission')

# seo = Pandrosus()
# seo.fetch('69c35332-6c8a-4f43-9eb2-e5e9cbe7f798', name='SeO2', mode='transmission')

# bunch = Kekropidai(name='Selenium standards')
# bunch.add(se)
# bunch.add(seo)

run_report('\t'+'Demeter')
from BMM.demeter import run_hephaestus

run_report('\t'+'machine learning and data evaluation')
from BMM.ml import BMMDataEvaluation
clf = BMMDataEvaluation()
    
    
run_report('\t'+'telemetry')
from BMM.telemetry import BMMTelemetry
tele = BMMTelemetry()


if BMMuser.element is None:
     try:
          BMMuser.element = rkvs.get('BMM:pds:element').decode('utf-8')
     except:
          pass
     try:
          BMMuser.edge    = rkvs.get('BMM:pds:edge').decode('utf-8')
     except:
          pass


#########################################################################
# ______ _____ _   _ _____ _____ _   _ _____ _   _ _____   _   _______  #
# |  ___|_   _| \ | |_   _/  ___| | | |_   _| \ | |  __ \ | | | | ___ \ #
# | |_    | | |  \| | | | \ `--.| |_| | | | |  \| | |  \/ | | | | |_/ / #
# |  _|   | | | . ` | | |  `--. \  _  | | | | . ` | | __  | | | |  __/  #
# | |    _| |_| |\  |_| |_/\__/ / | | |_| |_| |\  | |_\ \ | |_| | |     #
# \_|    \___/\_| \_/\___/\____/\_| |_/\___/\_| \_/\____/  \___/\_|     #
#########################################################################
                                                                    
      
run_report('\t'+'final setup: Xspress3')
from BMM.user_ns.dwelltime import with_xspress3, use_7element, use_4element, use_1element
from BMM.user_ns.detectors import xs4, xs1, xs7, xs
if BMMuser.element is not None and with_xspress3 is True: # make sure Xspress3 is configured to measure from the correct ROI
    if xs4 is not None and use_4element:
        BMMuser.verify_roi(xs4,  BMMuser.element, BMMuser.edge, tab='\t\t\t')
    if xs1 is not None and use_1element:
        BMMuser.verify_roi(xs1, BMMuser.element, BMMuser.edge, tab='\t\t\t')
    if xs7 is not None and use_7element:
        BMMuser.verify_roi(xs7, BMMuser.element, BMMuser.edge, tab='\t\t\t')
    #show_edges()

run_report('\t'+'final setup: cameras')
from BMM.user_ns.detectors import xascam, xrdcam, anacam
xascam._root = os.path.join(BMMuser.folder, 'snapshots')
xrdcam._root = os.path.join(BMMuser.folder, 'snapshots')
anacam._root = os.path.join(BMMuser.folder, 'snapshots')

     
run_report('\t'+'checking motor connections')
from BMM.edge import all_connected
if all_connected(True) is False:
     print(error_msg('Ophyd connection failure (testing main PDS motors)'))
     print(error_msg('You likely have to restart bsui.'))

run_report('\t'+'data folders and logging')
from BMM.user_ns.base import startup_dir
from BMM.user_ns.instruments import wmb, lmb, gmb, lsmb
if wmb  is not None: wmb.folder  = BMMuser.folder     # single or double wheel
if gawheel is not None:
    gawheel.folder = BMMuser.folder  # glancing angle stage
    gawheel.description = 'the glancing angle stage'
if lmb  is not None: lmb.folder  = BMMuser.folder     # Linkam stage
if lsmb is not None: lsmb.folder = BMMuser.folder     # LakeShore 331 temperature controller
if gmb  is not None: gmb.folder  = BMMuser.folder     # generic motor grid

run_report('\t'+'CMS experiment')
from BMM.agent_plans import CMS_driven_measurement


from BMM.logging import BMM_msg_hook
user_ns['RE'].msg_hook = BMM_msg_hook

def measuring(element, edge=None):
    BMMuser.element = element
    rkvs.set('BMM:pds:element', element)
    if edge is not None:
        BMMuser.edge = edge
        rkvs.set('BMM:pds:edge', edge)
    if use_7element:
        xs7.reset_rois()
    if use_4element:
        xs4.reset_rois()
    if use_1element:
        xs1.reset_rois()
    show_edges()


def xrf_measurement(stub=None, timestamp=True, post=False, add=False):
    '''Make a proper measurement of an XRF spectrum with the chosen xs
    detector.  Unlike the %xrf magic, this is made through the run
    engine and stored properly in the database.

    An image will be displayed and saved in the proposal XRF folder,
    as will and XDI-style ASCII data file.  The record will not,
    however, have the full XDI dictionary in the start document.

    '''
    if timestamp is True:
        xrffile = f'{stub}_{ahora}.xrf'
        xrfsnap = f'{stub}_{ahora}.png'
    else:
        xrffile = f'{stub}.xrf'
        xrfsnap = f'{stub}.png'
        
    yield from mv(xs.total_points, 1)
    yield from mv(xs.cam.acquire_time, 1)
    uid = yield from count([xs], 1, md = {'plan_name' : 'count xafs_metadata XRF'})

    kafka_message({'xrf' : 'plot',
                   'uid' : uid,
                   'add' : add,
                   'filename' : xrfsnap,
                   'post' : post, })
    kafka_message({'xrf' : 'write',
                   'uid' : uid,
                   'filename' : xrffile, })

def examine_diagnostics():
    CHECK = '\u2714'
    TAB = '\t\t\t'

    print(verbosebold_msg(f'\t\tverifying positions of diagnostics ...'))
    things = {'dm1_filters1' : ['DM1 filter ladder #1',   55],
              'dm1_filters2' : ['DM1 filter ladder #2',   55],
              'dm2_fs'       : ['DM2 fluorescent screen', 62],
              'dm3_fs'       : ['DM3 fluorescent screen', 55],
              'dm3_foils'    : ['DM3 foil ladder',        35],
              }

    for k in things.keys():
        if 'SynAxis' in f'{user_ns[k]}':
            print(disconnected_msg(f'{TAB}{things[k][0]} is not connected.'))
        elif user_ns[k].hocpl.get() == 1:
            if user_ns[k].position < things[k][1]-2: ## out of beam is 67
                print(error_msg(f'{TAB}{things[k][0]} is not out of the beam.'))
            else:
                print(f'{TAB}{things[k][0]} {CHECK}')
        elif 'filter' in k:
            print(whisper(f'{TAB}{things[k][0]} is not homed, but that\'s expected.'))
        else:
            print(whisper(f'{TAB}{things[k][0]} is not homed (which is ok, BR 5/5/23).'))
            #print(error_msg(f'{TAB}{things[k][0]} is not homed.'))

        
    if 'SynAxis' in f'{user_ns["dm3_bpm"]}':
        print(disconnected_msg(f'{TAB}DM3 BPM is not connected.'))
    elif user_ns['dm3_foils'].hocpl.get() == 1:
        if abs(user_ns['dm3_bpm'].position - 5.51) > 2: ## out of beam is 5.5112
            print(error_msg(f'{TAB}DM3 BPM is not out of the beam.'))
        else:
            print(f'{TAB}DM3 BPM {CHECK}')
    else:
        print(whisper(f'{TAB}DM3 BPM is not homed (which is ok, BR 5/5/23).'))
        #print(error_msg(f'{TAB}DM3 BPM is not homed.'))

            
def check_for_synaxis():
    '''A disconnected motor (due to IOC or controller not running) will be
    defined as a SynAxis. This does a test for that situation and
    reports about it at startup.  It also sets BMMuser.syns to True so
    things like motor_status() behave non-disastrously.
    '''
    BMMuser.syns = False
    remove_slits2, remove_slits3 = False, False
    syns = []
    for m in mcs8_motors+xafs_motors:
        if 'SynAxis' in f'{m}':
            syns.append(m.name)
            if m in user_ns['sd'].baseline:  # check is this is in the baseline, if so remove it
                i = user_ns['sd'].baseline.index(m)
                user_ns['sd'].baseline.pop(i)
            if 'dm3_slits' in m.name:
                remove_slits3 = True
            if 'dm2_slits' in m.name:
                remove_slits3 = True
                
    ## must handle slits specially due to inconsistent naming.  I blame Bruce!
    if remove_slits2 is True:
        sl2 = user_ns['slits3']
        for x in (sl2.bottom, sl2.top, sl2.inboard, sl2.outboard, sl2.vsize, sl2.vcenter, sl2.hsize, sl2.hcenter):
            i = user_ns['sd'].baseline.index(x)
            user_ns['sd'].baseline.pop(i)
    if remove_slits3 is True:
        sl3 = user_ns['slits3']
        for x in (sl3.bottom, sl3.top, sl3.inboard, sl3.outboard, sl3.vsize, sl3.vcenter, sl3.hsize, sl3.hcenter):
            i = user_ns['sd'].baseline.index(x)
            user_ns['sd'].baseline.pop(i)
            
        
    if len(syns) > 0:
        BMMuser.syns = True
        text = 'The following are disconnected & defined as simulated motors:\n\n'
        text += '\n'.join(disconnected_msg(x) for x in textwrap.wrap(', '.join(syns))) + '\n\n'
        text += 'This allows bsui to operate normally, but do not expect anything\n'
        text += 'involving those motors to work correctly.\n'
        text += whisper('(This likely means that an IOC or a motor controller (or both) are off.)\n')
        text += whisper('(Those motors have been removed from sd.baseline.)')
        boxedtext('Disconnected motors', text, 'red', width=74)
check_for_synaxis()
examine_diagnostics()
from BMM.workspace import check_instruments
check_instruments(user_ns['linkam'], user_ns['lakeshore'], user_ns['xs'])
run_report('\t  '+'calling resting_state')
resting_state()

if not is_re_worker_active():
    run_report('\t  '+'establishing local logger')
    from bluesky.log import config_bluesky_logging
    config_bluesky_logging(file='/home/xf06bm/logs/bluesky.log', level='DEBUG')


try:
    from bluesky_widgets.utils.streaming import stream_documents_into_runs
    from bluesky_widgets.models.plot_builders import Lines
    from bluesky_widgets.qt.figures import QtFigure
    # model = Lines("xafs_y", ["I0"], max_runs=1)
    # view = QtFigure(model.figure)
    # view.show()
except:
    pass


#import logging
#debug_monitor_log = logging.getLogger('ophyd.event_dispatcher')
#debug_monitor_log.addHandler(logging.StreamHandler(stream=sys.stdout))

from BMM.agent_plans import (agent_driven_nap, agent_move_and_measure, agent_move_motor, agent_change_edge,
                             agent_measure_single_edge)
