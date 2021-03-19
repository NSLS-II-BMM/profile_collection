run_report(__file__, text='instrument definitions')

## http://patorjk.com/software/taag/#p=display&f=Doom&t=MIRRORS

#################################################
# ___  ____________________ ___________  _____  #
# |  \/  |_   _| ___ \ ___ \  _  | ___ \/  ___| #
# | .  . | | | | |_/ / |_/ / | | | |_/ /\ `--.  #
# | |\/| | | | |    /|    /| | | |    /  `--. \ #
# | |  | |_| |_| |\ \| |\ \\ \_/ / |\ \ /\__/ / #
# \_|  |_/\___/\_| \_\_| \_|\___/\_| \_|\____/  #
#################################################


run_report('\tmirrors')
from BMM.motors import XAFSEpicsMotor, Mirrors, XAFSTable, GonioTable

## harmonic rejection mirror
m3_yu     = XAFSEpicsMotor('XF:06BMA-OP{Mir:M3-Ax:YU}Mtr',   name='m3_yu')
m3_ydo    = XAFSEpicsMotor('XF:06BMA-OP{Mir:M3-Ax:YDO}Mtr',  name='m3_ydo')
m3_ydi    = XAFSEpicsMotor('XF:06BMA-OP{Mir:M3-Ax:YDI}Mtr',  name='m3_ydi')
m3_xu     = XAFSEpicsMotor('XF:06BMA-OP{Mir:M3-Ax:XU}Mtr',   name='m3_xu')
m3_xd     = XAFSEpicsMotor('XF:06BMA-OP{Mir:M3-Ax:XD}Mtr',   name='m3_xd')
mcs8_motors.extend([m3_yu, m3_ydo, m3_ydi, m3_xu, m3_xd])



m1 = Mirrors('XF:06BM-OP{Mir:M1-Ax:',  name='m1', mirror_length=556,  mirror_width=240)
m1.vertical._limits = (-5.0, 5.0)
m1.lateral._limits  = (-5.0, 5.0)
m1.pitch._limits    = (-5.0, 5.0)
m1.roll._limits     = (-5.0, 5.0)
m1.yaw._limits      = (-5.0, 5.0)

m2 = Mirrors('XF:06BMA-OP{Mir:M2-Ax:', name='m2', mirror_length=1288, mirror_width=240)
m2.vertical._limits = (-6.0, 8.0)
m2.lateral._limits  = (-2, 2)
m2.pitch._limits    = (-0.5, 5.0)
m2.roll._limits     = (-2, 2)
m2.yaw._limits      = (-1, 2)

m3 = Mirrors('XF:06BMA-OP{Mir:M3-Ax:', name='m3', mirror_length=667,  mirror_width=240)
m3.vertical._limits = (-11, 1)
m3.lateral._limits  = (-16, 16)
m3.pitch._limits    = (-6, 6)
m3.roll._limits     = (-2, 2)
m3.yaw._limits      = (-1, 1)


def kill_mirror_jacks():
    yield from m2.kill_jacks()
    yield from m3.kill_jacks()


xt = xafs_table = XAFSTable('XF:06BMA-BI{XAFS-Ax:Tbl_', name='xafs_table', mirror_length=1160,  mirror_width=558)

gonio_table = GonioTable('XF:06BM-ES{SixC-Ax:Tbl_', name='gonio_table', mirror_length=1117.6,  mirror_width=711.12)

run_report('\tmirror trigonometry')
from BMM.mirror_trigonometry import move_m2, move_m3


###################################
#  _____ _     _____ _____ _____  #
# /  ___| |   |_   _|_   _/  ___| #
# \ `--.| |     | |   | | \ `--.  #
#  `--. \ |     | |   | |  `--. \ #
# /\__/ / |_____| |_  | | /\__/ / #
# \____/\_____/\___/  \_/ \____/  #
###################################
                               

run_report('\tslits')
from BMM.slits import Slits, GonioSlits, recover_slits2, recover_slits3

sl = slits3 = Slits('XF:06BM-BI{Slt:02-Ax:',  name='slits3')
slits3.nominal = [7.0, 1.0, 0.0, 0.0]
slits2 = Slits('XF:06BMA-OP{Slt:01-Ax:',  name='slits2')
slits2.nominal = [18.0, 1.1, 0.0, 0.6]
slits2.top.user_offset.put(-0.038)
slits2.bottom.user_offset.put(0.264)

        
slitsg = GonioSlits('XF:06BM-ES{SixC-Ax:Slt1_',  name='slitsg')

# We need to pass this device to plans executed in RE Worker, therefore we need the distinct name for it
slits3_hsize = slits3.hsize


#####################################
#  _    _ _   _  _____ _____ _      #
# | |  | | | | ||  ___|  ___| |     #
# | |  | | |_| || |__ | |__ | |     #
# | |/\| |  _  ||  __||  __|| |     #
# \  /\  / | | || |___| |___| |____ #
#  \/  \/\_| |_/\____/\____/\_____/ #
#####################################
                                 
                                 
run_report('\tsample wheels')
from BMM.wheel import WheelMotor, WheelMacroBuilder, reference, show_reference_wheel

xafs_wheel = xafs_rotb  = WheelMotor('XF:06BMA-BI{XAFS-Ax:RotB}Mtr',  name='xafs_wheel')
xafs_wheel.slotone = -30        # the angular position of slot #1
xafs_wheel.user_offset.put(-2.079)
slot = xafs_wheel.set_slot

xafs_ref = WheelMotor('XF:06BMA-BI{XAFS-Ax:Ref}Mtr',  name='xafs_ref')
xafs_ref.slotone = 0        # the angular position of slot #1

#                    1     2     3     4     5     6     7     8     9     10    11    12
xafs_ref.content = ['Gd', 'Ti', 'V',  'Cr', 'Mn', 'Fe', 'Co', 'Ni', 'Cu', 'Zn', 'Ga', 'Ge',
                    'As', 'Se', 'Br', 'Zr', 'Nb', 'Mo', 'Pt', 'Au', 'Pb', 'Bi', 'Ce', None]

        
## reference foil wheel will be something like this:
# xafs_wheel.content = ['Ti', 'V',  'Cr', 'Mn', 'Fe', 'Co', 'Ni', 'Cu', 'Zn', 'Ga', 'Ge', 'As',
#                       'Se', 'Br', 'Sr', 'Y',  'Nb', 'Mo', 'Hf', 'W',  'Re'  'Pt', 'Au', 'Pb']
#
# others: Zr, Rb, Ta, Hg, Ru, Bi, Cs, Ba, La, Ce to Lu (14)
#
# too low (H-Sc): 21
# noble gases: 4
# Rh to I: 9
# radioactive (Tc, Po, At, Fr - U): 9 (Tc, U, and Th could be part of the experiment)
#
# available = 47, unavailable = 45
#
# then,
#   try:
#      sl = xafs_wheel.content.index[elem] + 1
#      yield from reference_slot(sl)
#   except:
#      # don't move reference wheel


def setup_wheel():
    yield from mv(xafs_x, -119.7, xafs_y, 112.1, xafs_wheel, 0)
    
        
wmb = WheelMacroBuilder()
#xlsx = wmb.spreadsheet




######################################################################################
# ______ _____ _____ _____ _____ _____ ___________  ___  ________ _   _ _   _ _____  #
# |  _  \  ___|_   _|  ___/  __ \_   _|  _  | ___ \ |  \/  |  _  | | | | \ | |_   _| #
# | | | | |__   | | | |__ | /  \/ | | | | | | |_/ / | .  . | | | | | | |  \| | | |   #
# | | | |  __|  | | |  __|| |     | | | | | |    /  | |\/| | | | | | | | . ` | | |   #
# | |/ /| |___  | | | |___| \__/\ | | \ \_/ / |\ \  | |  | \ \_/ / |_| | |\  | | |   #
# |___/ \____/  \_/ \____/ \____/ \_/  \___/\_| \_| \_|  |_/\___/ \___/\_| \_/ \_/   #
######################################################################################
                                                                                  
run_report('\tdetector mount')
from BMM.detector_mount import DetectorMount
detx = DetectorMount()




###########################################################
#   ___  _____ _____ _   _  ___ _____ ___________  _____  #
#  / _ \/  __ \_   _| | | |/ _ \_   _|  _  | ___ \/  ___| #
# / /_\ \ /  \/ | | | | | / /_\ \| | | | | | |_/ /\ `--.  #
# |  _  | |     | | | | | |  _  || | | | | |    /  `--. \ #
# | | | | \__/\ | | | |_| | | | || | \ \_/ / |\ \ /\__/ / #
# \_| |_/\____/ \_/  \___/\_| |_/\_/  \___/\_| \_|\____/  #
###########################################################

run_report('\tactuators')
from BMM.actuators import BMPS_Shutter, IDPS_Shutter, EPS_Shutter, Spinner

try:
    bmps = BMPS_Shutter('SR:C06-EPS{PLC:1}', name='BMPS')
except:
    pass

try:
    idps = IDPS_Shutter('SR:C06-EPS{PLC:1}', name = 'IDPS')
except:
    pass


sha = EPS_Shutter('XF:06BM-PPS{Sh:FE}', name = 'Front-End Shutter')
sha.shutter_type = 'FE'
sha.openval  = 0
sha.closeval = 1
shb = EPS_Shutter('XF:06BM-PPS{Sh:A}', name = 'Photon Shutter')
shb.shutter_type = 'PH'
shb.openval  = 0
shb.closeval = 1

fs1 = EPS_Shutter('XF:06BMA-OP{FS:1}', name = 'FS1')
fs1.shutter_type = 'FS'
fs1.openval  = 1
fs1.closeval = 0


fan = Spinner('XF:06BM-EPS{Fan}', name = 'spinner')




###############################################
# ______ _____ _    _____ ___________  _____  #
# |  ___|_   _| |  |_   _|  ___| ___ \/  ___| #
# | |_    | | | |    | | | |__ | |_/ /\ `--.  #
# |  _|   | | | |    | | |  __||    /  `--. \ #
# | |    _| |_| |____| | | |___| |\ \ /\__/ / #
# \_|    \___/\_____/\_/ \____/\_| \_|\____/  #
###############################################


run_report('\tfilters')
from BMM.attenuators import attenuator, filter_state, set_filters

filter1 = attenuator()
filter1.motor = dm1_filters1
filter2 = attenuator()
filter2.motor = dm1_filters2



###############################################################################
# ____________ _____ _   _ _____      _____ _   _______  _______________  ___ #
# |  ___| ___ \  _  | \ | |_   _|    |  ___| \ | |  _  \ | ___ \ ___ \  \/  | #
# | |_  | |_/ / | | |  \| | | |______| |__ |  \| | | | | | |_/ / |_/ / .  . | #
# |  _| |    /| | | | . ` | | |______|  __|| . ` | | | | | ___ \  __/| |\/| | #
# | |   | |\ \\ \_/ / |\  | | |      | |___| |\  | |/ /  | |_/ / |   | |  | | #
# \_|   \_| \_|\___/\_| \_/ \_/      \____/\_| \_/___/   \____/\_|   \_|  |_/ #
###############################################################################

                                                                           
run_report('\tfront-end beam position monitor')
from BMM.frontend import FEBPM

try:
    bpm_upstream   = FEBPM('SR:C06-BI{BPM:4}Pos:', name='bpm_upstream')
    bpm_downstream = FEBPM('SR:C06-BI{BPM:5}Pos:', name='bpm_downstream')
except:
    pass

def read_bpms():
    return(bpm_upstream.x.get(), bpm_upstream.y.get(), bpm_downstream.x.get(), bpm_downstream.y.get())





#########################################################################################
# ___  ___  ___  _____ ______ _____  ______ _   _ _____ _    ______ ___________  _____  #
# |  \/  | / _ \/  __ \| ___ \  _  | | ___ \ | | |_   _| |   |  _  \  ___| ___ \/  ___| #
# | .  . |/ /_\ \ /  \/| |_/ / | | | | |_/ / | | | | | | |   | | | | |__ | |_/ /\ `--.  #
# | |\/| ||  _  | |    |    /| | | | | ___ \ | | | | | | |   | | | |  __||    /  `--. \ #
# | |  | || | | | \__/\| |\ \\ \_/ / | |_/ / |_| |_| |_| |___| |/ /| |___| |\ \ /\__/ / #
# \_|  |_/\_| |_/\____/\_| \_|\___/  \____/ \___/ \___/\_____/___/ \____/\_| \_|\____/  #
#########################################################################################
                                                                                     

from BMM.functions import present_options, bold_msg
from openpyxl import load_workbook
def xlsx():
    '''Prompt for a macro building spreadsheet for any instrument. Use the
    content of cell B1 to direct this spreadsheet to the correct builder.

    if cell B1 is "Glancing angle" --> build a glancing angle macro

    if cell B1 is "Sample wheel" --> build a sample wheel macro

    if cell B1 is empty --> build a sample wheel macro

    '''
    spreadsheet = present_options('xlsx')
    if spreadsheet is None:
        print(error_msg('No spreadsheet specified!'))
        return None
    #spreadsheet = os.path.join(BMMuser.folder, spreadsheet)
    wb = load_workbook(os.path.join(BMMuser.folder, spreadsheet), read_only=True);
    ws = wb.active
    instrument = str(ws['B1'].value).lower()
    if instrument == 'glancing angle':
        print(bold_msg('This is a glancing angle spreadsheet'))
        pinwheel.spreadsheet(spreadsheet)
    else:
        print(bold_msg('This is a sample wheel spreadsheet'))
        wmb.spreadsheet(spreadsheet)

####################################################################################
#  _   _______ _      _       _____  _    _ _____ _____ _____  _   _  _____ _____  #
# | | / /_   _| |    | |     /  ___|| |  | |_   _|_   _/  __ \| | | ||  ___/  ___| #
# | |/ /  | | | |    | |     \ `--. | |  | | | |   | | | /  \/| |_| || |__ \ `--.  #
# |    \  | | | |    | |      `--. \| |/\| | | |   | | | |    |  _  ||  __| `--. \ #
# | |\  \_| |_| |____| |____ /\__/ /\  /\  /_| |_  | | | \__/\| | | || |___/\__/ / #
# \_| \_/\___/\_____/\_____/ \____/  \/  \/ \___/  \_/  \____/\_| |_/\____/\____/  #
####################################################################################
                                                                                

run_report('\tamplifier kill switches')
from BMM.killswitch import KillSwitch
ks = KillSwitch('XF:06BMB-CT{DIODE-Local:4}', name='amplifier kill switches')
