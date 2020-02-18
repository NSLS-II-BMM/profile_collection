
import sys, os.path, re
#import pprint
#pp = pprint.PrettyPrinter(indent=4)
from openpyxl import load_workbook
import configparser

run_report(__file__)


class WheelMotor(EndStationEpicsMotor):
    '''Motor class for BMM sample wheels.
    
    These wheels have 24 slots, spaced 15 degrees apart.

    current_slot():
       return the current slot number, even if the wheel has rotated many times

    set_slot(n):
       move to the given slot number, taking care to go the shorter way
    '''
    def current_slot(self, value=None):
        '''Return the current slot number for a sample wheel.'''
        if value is not None:
            angle = round(value)
        else:
            angle = round(self.user_readback.value)
        this = round((-1*self.slotone-15+angle) / (-15)) % 24
        if this == 0: this = 24
        return this

    def reset(self):
        '''Return a sample wheel to slot 1'''
        yield from mv(self, self.slotone)
        report('Returned to sample wheel slot 1 at %d degrees' % self.slotone, 'bold')

    def angle_from_current(self, n):
        '''Return the angle from the current position to the target'''
        if type(n) is not int:
            print(error_msg('Slots numbers must be integers between 1 and 24 (argument was %s)' % type(n)))
            return(0)
        if n<1 or n>24:
            print(error_msg('Slots are numbered from 1 to 24'))
            return(0)

        current = self.current_slot()
        distance = n - current
        if distance > 12:
            distance = distance - 24
        elif distance < -12:
            distance = 24 + distance
        angle = -15*distance
        return angle
        
    def set_slot(self, n):
        '''Move to a numbered slot on a sample wheel.'''
        if type(n) is not int:
            print(error_msg('Slots numbers must be integers between 1 and 24 (argument was %s)' % type(n)))
            return(yield from null())
        if n<1 or n>24:
            print(error_msg('Slots are numbered from 1 to 24'))
            return(yield from null())
        angle = self.angle_from_current(n)
        yield from mvr(self, angle)
        report('Arrived at sample wheel slot %d' % n, 'bold')

    def slot_number(self, target=None):
        try:
            target = target.capitalize()
            slot = xafs_ref.content.index(target) + 1
            return slot
        except:
            return self.current_slot()
                
    def position_of_slot(self, target):
        if type(target) is int:
            angle = self.angle_from_current(target)
        elif type(target) is str:
            target = self.slot_number(target.capitalize())
            angle = self.angle_from_current(target)
        return(xafs_ref.user_readback.value+angle)

    def recenter(self):
        here = self.user_readback.value
        center = round(here / 15) * 15
        yield from mv(self, center)
        print(whisper('recentered %s to %.1f' % (self.name, center)))
    

xafs_wheel = xafs_rotb  = WheelMotor('XF:06BMA-BI{XAFS-Ax:RotB}Mtr',  name='xafs_wheel')
xafs_wheel.slotone = -30        # the angular position of slot #1
xafs_wheel.user_offset.put(-2.079)
slot = xafs_wheel.set_slot

xafs_ref = WheelMotor('XF:06BMA-BI{XAFS-Ax:Ref}Mtr',  name='xafs_ref')
xafs_ref.slotone = 0        # the angular position of slot #1

#                    1     2     3     4     5     6     7     8     9     10    11    12
xafs_ref.content = [None, 'Ti', 'V',  'Cr', 'Mn', 'Fe', 'Co', 'Ni', 'Cu', 'Zn', 'Ga', 'Ge',
                    'As', 'Se', 'Br', 'Zr', 'Nb', 'Mo', 'Pt', 'Au', 'Pb', None, None, None]

def reference(target=None):
    if target is None:
        print('Not moving reference wheel.')
        return(yield from null())
    if type(target) is int:
        if target < 1 or target > 24:
            print('An integer reference target must be between 1 and 24 (%d)' % target)
            return(yield from null())
        else:
            yield from xafs_ref.set_slot(target)
            return
    try:
        target = target.capitalize()
        slot = xafs_ref.content.index(target) + 1
        yield from xafs_ref.set_slot(slot)
    except:
        print('Element %s is not on the reference wheel.' % target)
        

def show_reference_wheel():
    wheel = xafs_ref.content.copy()
    this  = xafs_ref.current_slot() - 1
    #wheel[this] = go_msg(wheel[this])
    text = 'Foil wheel:\n'
    text += bold_msg('    1      2      3      4      5      6      7      8      9     10     11     12\n')
    text += ' '
    for i in range(12):
        if i==this:
            text += go_msg('%4.4s' % str(wheel[i])) + '   '
        else:
            text += '%4.4s' % str(wheel[i]) + '   '
    text += '\n'
    text += bold_msg('   13     14     15     16     17     18     19     20     21     22     23     24\n')
    text += ' '
    for i in range(12, 24):
        if i==this:
            text += go_msg('%4.4s' % str(wheel[i])) + '   '
        else:
            text += '%4.4s' % str(wheel[i]) + '   '
    text += '\n'
    return(text)
        
## reference foil wheel will be something like this:
# xafs_wheel.content = ['Ti', 'V',  'Cr', 'Mn', 'Fe', 'Co', 'Ni', 'Cu', 'Zn', 'Ga', 'Ge', 'As',
#                       'Se', 'Br', 'Sr', 'Y',  'Nb', 'Mo', 'Hf', 'W',  'Re'  'Pt', 'Au', 'Pb']
#
# others: Zr, Rb, Ta, Hg, Ru, Bi, Cs, Ba, La, Ce to Lu (14)
#
# too low (H-Sc): 21
# noble gases: 4
# Rh to I: 9
# radioactive (Tc, Po, At, Fr - U): 9 (Tc, U, and Th could be part of the experiment)
#
# available = 47, unavailable = 45
#
# then,
#   try:
#      sl = xafs_wheel.content.index[elem] + 1
#      yield from reference_slot(sl)
#   except:
#      # don't move reference wheel


def setup_wheel():
    yield from mv(xafs_x, -119.7, xafs_y, 112.1, xafs_wheel, 0)
    


class WheelMacroBuilder():
    '''A class for parsing specially constructed spreadsheets and
    generating macros for measuring XAS on the BMM wheel.

    Example:
       mb = MacroBuilder()
       mb.spreadsheet('wheel1.xlsx')
       mb.write_macro()
    '''
    def __init__(self):
        self.basename     = None
        self.folder       = BMMuser.folder

        self.source       = None
        self.wb           = None
        self.ws           = None
        self.measurements = list()
        self.ini          = None
        self.macro        = None

        self.tab          = '        '
        self.content      = ''
        self.do_first_change = False
        self.has_e0_column   = False
        self.verbose         = False

    def spreadsheet(self, spreadsheet, energy=False):
        '''Convert a wheel macro spreadsheet to a BlueSky plan.

        To create a macro from a spreadsheet called "MySamples.xlsx"

            xlsx('MySamples')

        To specify a change_edge() command at the beginning of the macro:

            xlsx('MySamples', energy=True)

        '''
        if spreadsheet[-5:] != '.xlsx':
            spreadsheet = spreadsheet+'.xlsx'
        self.source   = os.path.join(self.folder, spreadsheet)
        self.basename = os.path.splitext(spreadsheet)[0]
        self.wb       = load_workbook(self.source, read_only=True);
        self.ws       = self.wb.active
        self.ini      = os.path.join(self.folder, self.basename+'.ini')
        self.tmpl     = os.path.join(os.getenv('HOME'), '.ipython', 'profile_collection', 'startup', 'wheelmacro.tmpl')
        self.macro    = os.path.join(self.folder, self.basename+'_macro.py')
        self.measurements = list()
        self.do_first_change = False
        if energy is True:
            self.do_first_change = True

        if self.ws['H5'].value == 'e0': # accommodate older xlsx files which have e0 values in column H
            self.has_e0_column = True
            
        self.read_spreadsheet()
        self.write_macro()

        
    def truefalse(self, value):
        '''Interpret certain strings from the spreadsheet as True/False'''
        if value is None:
            return True # self.measurements[0]['measure']
        if str(value).lower() == '=true()':
            return True
        elif str(value).lower() == 'true':
            return True
        elif str(value).lower() == 'yes':
            return True
        else:
            return False


    def ini_sanity(self, default):
        '''Sanity checks for the default line from the spreadsheet.

        1. experimenters is a string (BMMuser.name)
        2. sample, prep, and comment are not empty strings (set to '...')
        3. nscans is an integer (set to 1)
        4. start is an integer or "next"
        5. mode is string (set to 'transmission')
        6. element is an element (bail)
        7. edge is k, l1, l2, or l3 (bail)
        
        To do:
          * booleans are interpretable as booleans
          * focused is focused or unfocused
          * bounds, steps, times are sensible
          * x, y, slits are floats and sensible for the respective ranges of motion

        '''

        message = ''
        unrecoverable = False
        
        if default['experimenters'] is None or str(default['experimenters']).strip() == '':
            default['experimenters'] = BMMuser.name

        for k in ('sample', 'prep', 'comment'):
            if default[k] is None or str(default[k]).strip() == '':
                default[k] = '...'

        try:
            default['nscans'] = int(default['nscans'])
        except:
            default['nscans'] = 1

        try:
            default['start'] = int(default['start'])
        except:
            default['start'] = 'next'
            
        if default['mode'] is None or str(default['mode']).strip() == '':
            default['mode'] = 'transmission'

        if str(default['element']).capitalize() not in re.split('\s+', PERIODIC_TABLE): # see 06-periodic table 
            message += '\nDefault entry for element is not recognized.'
            unrecoverable = True

        if str(default['edge']).lower() not in ('k', 'l1', 'l2', 'l3'):
            message += '\nDefault entry for edge is not recognized.'
            unrecoverable = True

        # try:
        #     default['e0'] = float(default['e0'])
        # except:
        #     default['e0'] = edge_energy(default['element'], default['edge'])

        if unrecoverable:
            print(error_msg(message))
            default = None
        return default

        
    def write_macro(self):
        '''Write a macro paragraph for each sample described in the
        spreadsheet.  A paragraph consists of line to move to the
        correct wheel slot, a line to change the edge energy (if
        needed), a line to measure the XAFS using the correct set of
        control parameters, and a line to close plot windows after the
        scan.  

        Finally, write out the master INI and macro python files.
        '''
        element, edge, focus = (None, None, None)
        self.content = ''
        for m in self.measurements:

            #####################################################
            # all the reasons to skip a line in the spreadsheet #
            #####################################################
            if m['default'] is True:
                element = m['element']
                edge    = m['edge']
                continue
            if type(m['slot']) is not int:
                continue
            if type(m['filename']) is None:
                continue
            if  self.truefalse(m['measure']) is False:
                continue
            if m['nscans'] is not None and m['nscans'] < 1:
                continue

            #######################################
            # default element/edge(/focus) values #
            #######################################
            for k in ('element', 'edge'):
                if m[k] is None:
                    m[k] = self.measurements[0][k]

            ############################
            # sample and slit movement #
            ############################
            self.content += self.tab + 'yield from slot(%d)\n' % m['slot']
            if m['samplex'] is not None:
                self.content += self.tab + 'yield from mv(xafs_x, %d)\n' % m['samplex']
            if m['sampley'] is not None:
                self.content += self.tab + 'yield from mv(xafs_y, %d)\n' % m['sampley']
            if m['slitwidth'] is not None:
                self.content += self.tab + 'yield from mv(slits3.hsize, %.2f)\n' % m['slitwidth']

            
            ##########################
            # change edge, if needed #
            ##########################
            focus = False
            if m['focus'] == 'focused':
                focus = True
            if self.do_first_change is True:
                self.content += self.tab + 'yield from change_edge(\'%s\', edge=\'%s\', focus=%r)\n' % (m['element'], m['edge'], focus)
                self.do_first_change = False
                
            elif m['element'] != element or m['edge'] != edge: # focus...
                element = m['element']
                edge    = m['edge']
                self.content += self.tab + 'yield from change_edge(\'%s\', edge=\'%s\', focus=%r)\n' % (m['element'], m['edge'], focus)
            else:
                if self.verbose:
                    self.content += self.tab + '## staying at %s %s\n' % (m['element'], m['edge'])
                pass

            ######################################
            # measure XAFS, then close all plots #
            ######################################
            command = self.tab + 'yield from xafs(\'%s.ini\'' % self.basename
            for k in m.keys():
                ## skip cells with macro-building parameters that are not INI parameters
                if k in ('default', 'slot', 'focus', 'measure'):
                    continue
                ## skip the flags for now
                elif k in ('snapshots', 'htmlpage', 'usbstick', 'bothways', 'channelcut', 'ththth'):
                    continue
                elif k in ('samplex', 'sampley', 'slitwidth'):
                    continue
                ## skip element & edge if they are same as default
                elif k in ('element', 'edge'):
                    if m[k] == self.measurements[0][k]:
                        continue
                ## skip cells with only whitespace
                if type(m[k]) is str and len(m[k].strip()) == 0:
                    m[k] = None
                ## if a cell has data, put it in the argument list for xafs()
                if m[k] is not None:
                    if type(m[k]) is int:
                        command += ', %s=%d' % (k, m[k])
                    elif type(m[k]) is float:
                        command += ', %s=%.3f' % (k, m[k])
                    else:
                        command += ', %s=\'%s\'' % (k, m[k])
            command += ')\n'
            self.content += command
            self.content += self.tab + 'close_last_plot()\n\n'

        
        #################################
        # write out the master INI file #
        #################################
        config = configparser.ConfigParser()
        default = self.measurements[0].copy()
        for k in ('default', 'slot', 'measure', 'focus', 'samplex', 'sampley', 'slitwidth'): # things in the spreadsheet but not in the INI file
            default.pop(k, None)
        default['experimenters'] = self.ws['E1'].value # top line of xlsx file
        default = self.ini_sanity(default)
        if default is None:
            print(error_msg("Could not interpret %s as a wheel macro." % self.source))
            return
        config.read_dict({'scan': default})
        with open(self.ini, 'w') as configfile:
            config.write(configfile)
        print(whisper('Wrote default INI file: %s' % self.ini))

        ########################################################
        # write the full macro to a file and %run -i that file #
        ########################################################
        with open(self.tmpl) as f:
            text = f.readlines()
        fullmacro = ''.join(text).format(folder=self.folder, base=self.basename, content=self.content)
        o = open(self.macro, 'w')
        o.write(fullmacro)
        o.close()
        from IPython import get_ipython
        ipython = get_ipython()
        ipython.magic('run -i \'%s\'' % self.macro)
        print(whisper('Wrote macro file: %s' % self.macro))
            
        #######################################
        # explain to the user what to do next #
        #######################################
        print('\nYour new plan is called: ' + bold_msg('%s_macro' % self.basename))
        print('\nVerify: ' + bold_msg('%s_macro??' % self.basename))
        print('Dryrun: '   + bold_msg('RE(%s_macro(dryrun=True))' % self.basename))
        print('Run:    '   + bold_msg('RE(%s_macro())' % self.basename))

            
    def read_spreadsheet(self):
        '''Slurp up the content of the spreadsheet and write the default control file
        '''
        print('Reading spreadsheet: %s' % self.source)
        count = 0
        offset = 0
        if self.has_e0_column:  # deal with older xlsx that have e0 in column H
            offset = 1
        for row in self.ws.rows:
            count += 1
            if count < 6:
                continue
            defaultline = False
            if count == 6:
                defaultline = True
            self.measurements.append({'default' :   defaultline,
                                      'slot':       row[1].value,      # sample location
                                      'measure':    self.truefalse(row[2].value),  # filename and visualization
                                      'filename':   row[3].value,
                                      'nscans':     row[4].value,
                                      'start':      row[5].value,
                                      'mode':       row[6].value,
                                      #'e0':         row[7].value,      
                                      'element':    row[7+offset].value,      # energy range
                                      'edge':       row[8+offset].value,
                                      'focus':      row[9+offset].value,
                                      'sample':     row[10+offset].value,     # scan metadata
                                      'prep':       row[11+offset].value,
                                      'comment':    row[13+offset].value,
                                      'bounds':     row[13+offset].value,     # scan parameters
                                      'steps':      row[14+offset].value,
                                      'times':      row[15+offset].value,
                                      'samplex':    row[16+offset].value,     # other motors 
                                      'sampley':    row[17+offset].value,
                                      'slitwidth':  row[18+offset].value,
                                      'snapshots':  self.truefalse(row[19+offset].value), # flags
                                      'htmlpage':   self.truefalse(row[20+offset].value),
                                      'usbstick':   self.truefalse(row[21+offset].value),
                                      'bothways':   self.truefalse(row[22+offset].value),
                                      'channelcut': self.truefalse(row[23+offset].value),
                                      'ththth':     self.truefalse(row[24+offset].value),
            })
        #pp.pprint(self.measurements)
        
wmb = WheelMacroBuilder()
#wmb.do_first_change = True

xlsx = wmb.spreadsheet
